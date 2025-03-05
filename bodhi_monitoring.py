#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: ijeong-yeon
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

class Bodhi_monitoring:

    def __init__(self, name, df):
        """
        - Initialise the data collection monitoring pipeline class

        name: str, Name of the project
        indicators: list, List of the project indicators
        df: dataframe, Primary quantitative dataset
        """
        self.name = name
        self.df = df

    def setting(self, pilot_test_dates, enumerator_names, location, gender, age, disability, respondent_type = None):
        """
        - Data collection monitoring model set-up 

        pilot_test_dates: list, Dates on which the pilot test was conducted from the data
        enumerator_names: list, List of columns that have enumerators' names
        location: str, Name of the location column
        gender: str, Name of the gender column
        age: str, Name of the age column
        disability: list, List of all WG-SS columns
        respondent_type: str, Name of the respondent type column (where applicable)
        """
        df = self.df
        
        df['today'] = df['today'].dt.strftime('%Y-%m-%d')
        df = df.dropna(subset=['today'])
        if len(pilot_test_dates) != 0:
            for date in pilot_test_dates:
                df = df[df['today'] != date]
                
        if len(enumerator_names) != 1:
            df['Enumerator Name'] = df[[enumerator_names]].bfill(axis=1).iloc[:, 0]
            self.enumerator_name = "Enumerator Name"
        else: self.enumerator_name = enumerator_names[0]
                    
        bins = [0, 17, 24, 34, 44, 54, 64, float('inf')]
        labels = ['Below 18','18 - 24','25 - 34', '35 - 44', '45 - 54', '55 - 64', 'Above 65 years']
        df['Age Group'] = pd.cut(df[age], bins=bins, labels=labels, right=True)
        try:
            df['WG-Disability'] = ''
            def wg_ss(row, cols):
                values = row[cols]
                some_difficulty_count = (values == 'Some difficulty').sum()
                a_lot_of_difficulty = (values == 'A lot of difficulty').any() or (values == 'Cannot do at all').any()
                cannot_do_at_all = (values == 'Cannot do at all').any()
                if cannot_do_at_all:
                    return 'DISABILITY4'
                elif a_lot_of_difficulty:
                    return 'DISABILITY3'
                elif some_difficulty_count >= 2:
                    return 'DISABILITY2'
                elif some_difficulty_count >= 1:
                    return 'DISABILITY1'
                else:
                    return 'No_disability'
            df['WG-Disability'] = df.apply(lambda row: wg_ss(row, disability), axis=1)
            df['Disability'] = df['WG-Disability'].apply(lambda x: 'Disability' if x in ['DISABILITY4', 'DISABILITY3'] else 'No Disability')
        except Exception as e:
            print('New disability variable has not been created in this dataset')    
        self.disability = 'Disability'
        self.age_group = 'Age Group'
        self.location = location
        self.gender = gender
        if respondent_type != None:
            self.respondent_type = respondent_type
        else: self.respondent_type = None
        self.df = df
        print("The data collection monitoring pipeline has been established.")

    def run(self):
        df = self.df
        def adjust_column_width(sheet):
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        if self.respondent_type == None:
            df_0 = df[self.location].value_counts().reset_index(name='count')
            df_1 = df.groupby([self.location])[self.gender].value_counts().reset_index(name='count')
            df_2 = df.groupby([self.location])[self.age_group].value_counts().reset_index(name='count')
            df_3 = df.groupby([self.location])[self.disability].value_counts().reset_index(name='count')
            df_4 = df.groupby([self.location])['today'].value_counts().reset_index(name='count')
            df_pivot = df_4.pivot_table(index=[self.location], columns='today', values='count', fill_value=0)
            df_5 = df.groupby([self.location, self.enumerator_name, 'today'])[self.gender].value_counts().reset_index(name='count')
            df_pivot2 = df_5.pivot_table(index=[self.location, self.enumerator_name, self.gender], columns='today', values='count', fill_value=0)
            df_6 = df.groupby([self.enumerator_name])[self.location].value_counts().reset_index(name='count')
            df_pivot3 = df_6.pivot_table(index=[self.enumerator_name], columns=self.location, values='count', fill_value=0)
        else:
            df_0 = df.groupby([self.location])[self.respondent_type].value_counts().reset_index(name='count')
            df_1 = df.groupby([self.location, self.respondent_type])[self.gender].value_counts().reset_index(name='count')
            df_2 = df.groupby([self.location, self.respondent_type])[self.age_group].value_counts().reset_index(name='count')
            df_3 = df.groupby([self.location, self.respondent_type])[self.disability].value_counts().reset_index(name='count')
            df_4 = df.groupby([self.location, 'today'])[self.respondent_type].value_counts().reset_index(name='count')
            df_pivot = df_4.pivot_table(index=[self.location, self.respondent_type], columns='today', values='count', fill_value=0)
            df_5 = df.groupby([self.location, self.enumerator_name, self.respondent_type, 'today'])[self.gender].value_counts().reset_index(name='count')
            df_pivot2 = df_5.pivot_table(index=[self.location, self.enumerator_name, self.respondent_type, self.gender], columns='today', values='count', fill_value=0)
            df_6 = df.groupby([self.location, self.enumerator_name])[self.respondent_type].value_counts().reset_index(name='count')
            df_pivot3 = df_6.pivot_table(index=[self.location,self.enumerator_name], columns=self.respondent_type, values='count', fill_value=0)
    
        with pd.ExcelWriter(f"data/{self.name}_data_collection_monitoring.xlsx", engine="openpyxl") as writer:
            df_pivot.to_excel(writer, sheet_name="Overall", startrow=1)
            start_row = len(df_pivot) + 3
            df_0.to_excel(writer, sheet_name='Overall', startrow=start_row, index=False, header=True)
            df_pivot3.to_excel(writer, sheet_name="Overall_enumerators", startrow=1)
            df_pivot2.to_excel(writer, sheet_name="Daily_enumerators", startrow=1)
            
            start_row = 1
            df_1.to_excel(writer, sheet_name="Detail", startrow=start_row + 1, index=False)
            df_2.to_excel(writer, sheet_name="Detail", startrow=start_row + len(df_1) + 4, index=False)
            df_3.to_excel(writer, sheet_name="Detail", startrow=start_row + len(df_1) + len(df_2) + 7, index=False)
    
        book = load_workbook(f"data/{self.name}_data_collection_monitoring.xlsx")
    
        sheet = book["Overall"]
        sheet.cell(row=1, column=1, value="Overall data collection")

        sheet = book["Overall_enumerators"]
        sheet.cell(row=1, column=1, value="Overall data collection (Enumerators)")
    
        sheet = book["Daily_enumerators"]
        sheet.cell(row=1, column=1, value="Daily data collection by gender (Enumerators)")
    
        sheet = book["Detail"]
        sheet.cell(row=start_row, column=1, value="Overall data collection by gender")
        sheet.cell(row=start_row + len(df_1) + 4, column=1, value="Overall data collection by age group")
        sheet.cell(row=start_row + len(df_1) + len(df_2) + 7, column=1, value="Overall data collection by disability status")

        adjust_column_width(book['Overall'])
        adjust_column_width(book["Overall_enumerators"])
        adjust_column_width(book["Daily_enumerators"])
        adjust_column_width(book["Detail"])
    
        book.save(f"data/{self.name}_data_collection_monitoring.xlsx")
    
        print("Excel export completed successfully! (Please have a look at the data folder)")
        print(f"File name = {self.name}_data_collection_monitoring.xlsx")